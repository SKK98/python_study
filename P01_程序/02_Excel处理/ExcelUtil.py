import os
from openpyxl import Workbook, load_workbook


class ExcelUtils:

    def __init__(self):
        pass

    # 字典写成Excel的通用方法
    def write_dict_to_excel(dict_data, file_name):
        # 如果文件不存在，创建文件和工作表
        if not os.path.exists(file_name):
            workbook = Workbook()  # 创建工作簿
            worksheet = workbook.active  # 创建工作表
            # 写入表头
            header = list(dict_data.keys())  # 获取字典的键
            for col in range(len(header)):  # 遍历表头
                worksheet.cell(row=1, column=col + 1, value=header[col])  # 写入表头数据 从第一行第一列开始写入 行值和列值都是从1开始
        else:
            # 如果文件已存在，打开文件和工作表
            workbook = load_workbook(file_name)  # 打开工作簿
            worksheet = workbook.active  # 打开工作表

        # 获取已有数据的行数,确定新写入的起始行
        row = worksheet.max_row  # 获取行数

        # 写入数据
        for col, value in enumerate(dict_data.values()):  # 遍历字典的值
            worksheet.cell(row=row + 1, column=col + 1, value=value)  # 写入数据

        workbook.save(file_name)
