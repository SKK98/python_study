import re
import openpyxl

import xml.etree.ElementTree as ET
import openpyxl



def txt_to_excel(txt_file):
    # 读取txt内容
    with open(txt_file, 'r', encoding='utf-8') as f:
        text = f.read()

    # 使用正则表达式提取NewStoredProcedureCall标签内容
    pattern = re.compile(r'<NewStoredProcedureCall>(.*?)</NewStoredProcedureCall>', re.DOTALL)
    results = re.findall(pattern, text)

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'NewStoredProcedureCall'

    # 写入表头
    headers = ['ID', 'NUMCODE', 'NAME', 'SEXID', 'SEXNAME', 'IDNO',
               'STFTYPEID', 'STFTYPENAME', 'DESP', 'DEPTID', 'DEPTNAME',
               'DEPTCODE', 'TELEPHONE', 'POSTID', 'POSTNAME', 'STFTITLE',
               'INDATE', 'OUTDATE', 'ISWAS', 'HASRX', 'KZLDRUG1',
               'KZLDRUG2', 'KJSKD1', 'KJSKD2', 'KJSKD3']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header

    # 逐行写入数据
    for row, result in enumerate(results, 2):
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            match = re.search(r'<{}>(.*?)</{}>'.format(header, header), result)
            if match:
                cell.value = match.group(1)

    # 保存Excel
    wb.save('data.xlsx')
    pass



if __name__ == '__main__':
    txt_to_excel("data.txt")


